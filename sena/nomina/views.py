import os
from itertools import chain
import base64
from django.contrib.auth.hashers import make_password
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template.loader import render_to_string
from django.utils.text import slugify
from django.utils.dateparse import parse_date
from django.template.defaultfilters import date as date_filter
from django.core.mail import BadHeaderError, EmailMessage
from django.conf import settings
import datetime
import pdfkit
from django.db import IntegrityError, transaction
import datetime
import openpyxl

from django.utils.timezone import localdate
from openpyxl import Workbook
import xlwings as xw
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q, Max
from django.urls import reverse
from django.contrib import messages
from .models import *
from django.shortcuts import get_object_or_404
from xhtml2pdf import pisa
import io
from io import BytesIO
from django.db.models.functions import TruncMonth
from django.db.models import Count
from urllib.parse import urlencode
from django.utils import timezone


# Create your views here.

def login(request):
    if request.method == "POST":
        correo = request.POST.get("correo_login")
        contrasena = request.POST.get("contrasena_login")
        remember_me = request.POST.get("remember_me", False)

        try:
            q = Usuario.objects.get(correo=correo, contrasena=contrasena)
            messages.success(request, f"Bienvenido sr(a) {q.nombre}!!")
            datos = {
                "rol": q.rol,
                "nombre_rol": q.get_rol_display(),
                "nombre": f"{q.nombre} {q.apellido}",
                "foto": q.foto.url if q.foto else "/media/fotos/default.png",
                "id": q.id
            }
            request.session["logueo"] = datos

            # Manejo del "Recordar usuario"
            if remember_me:
                request.session.set_expiry(1209600)  # 2 semanas
            else:
                request.session.set_expiry(0)  # Cerrar la sesión al cerrar el navegador

            return HttpResponseRedirect(reverse("nomina:index"))
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario o contraseña no válidos...")
            return render(request, "nomina/login.html")
    else:
        if request.session.get("logueo", False):
            return HttpResponseRedirect(reverse("nomina:index"))
        else:
            return render(request, "nomina/login.html")


def logout(request):
    try:
        del request.session["logueo"]
        messages.success(request, "Cesión cerrada correctamente")
    except Exception as e:
        messages.error(request, f"Error: {e}")
    return HttpResponseRedirect(reverse("nomina:login"))


def landing_page(request):
    if request.session.get("logueo", False):
        return render(request, "nomina/index.html")
    else:
        return render(request, 'nomina/landing-page.html')


def index(request):
    if request.session.get("logueo", False):
        # Obtén el porcentaje de incapacidad de cualquiera de las novedades
        incapacidad_porcentaje = Novedad.objects.values_list('incapacidad_porcentaje', flat=True).first()

        # Si no hay novedades, el porcentaje es 100% por defecto
        if incapacidad_porcentaje is None:
            incapacidad_porcentaje = 1.0  # Porcentaje por defecto 100% (como flotante)

        context = {
            'incapacidad_porcentaje': incapacidad_porcentaje * 100,  # Lo multiplicas por 100 para mostrarlo correctamente
        }

        return render(request, "nomina/index.html", context)
    else:
        return render(request, 'nomina/login.html')


def registrarse(request):
    roles = Usuario.ROLES
    context = {
        'ROLES': roles,
    }
    return render(request, 'nomina/registrarse.html', context)


def crear_admin(request):
    if request.method == "POST":
        cedula = request.POST.get("cedula")
        nombre = request.POST.get('nombre')
        apellido = request.POST.get("apellido")
        correo = request.POST.get('correo')
        salario = request.POST.get('salario')
        contrasena = cedula
        rol = 1
        foto = request.FILES.get("foto")

        try:
            usu = Usuario(
                cedula=cedula,
                nombre=nombre,
                apellido=apellido,
                correo=correo,
                salario=salario,
                contrasena=contrasena,
                rol=rol,
                foto=foto,
            )
            usu.save()
            messages.success(request, "Guardado correctamente!!")
            try:
                destinatario = usu.correo

                mensaje = f"""
                    <h1 style='color:blue;'>¡Ya eres parte de nuestra familia <strong>PRESS</strong>!</h1>
                    <p>Tu registro fue exitoso. Ya puedes acceder a todas nuestras funciones ingresando con tu correo electrónico y esta contraseña: {usu.contrasena}. Puedes cambiarla cuando desees.</p>
                    <a href='https://senapress.pythonanywhere.com'>Ingresa desde este link</a>
                """

                try:
                    msg = EmailMessage("PRESS", mensaje, settings.EMAIL_HOST_USER, [destinatario])
                    msg.content_subtype = "html"  # Habilitar html
                    msg.send()
                except BadHeaderError:
                    return HttpResponse("Invalid header found.")
                except Exception as e:
                    return HttpResponse(f"Error: {e}")

            except Exception as e:
                print(f"{e}")

        except Exception as e:
            messages.error(request, f"Error. {e}")

        return HttpResponseRedirect(reverse("nomina:index", args=()))

    else:
        messages.warning(request, "No se enviarion datos...")
        return HttpResponseRedirect(reverse("nomina:usuario_formulario", args=()))


def convert_html_to_pdf(source_html):
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(source_html.encode("UTF-8")), result)
    if not pdf.err:
        return result.getvalue()
    return None


def recibo_view(request, nomina_id):
    if request.session.get("logueo", False):
        usuario_id = request.session["logueo"]["id"]
        usuario = Usuario.objects.get(id=usuario_id)
        nomina = get_object_or_404(Nomina, id=nomina_id)
        he_diurnas = nomina.novedad.horas_extras_diurnas or 0
        he_diurnas_df = nomina.novedad.horas_extras_diurnas_dom_fes or 0
        he_nocturnas = nomina.novedad.horas_extras_nocturnas or 0
        he_nocturnas_df = nomina.novedad.horas_extras_nocturnas_dom_fes or 0
        hr_diurno_df = nomina.novedad.horas_recargo_diurno_dom_fes or 0
        hr_nocturo = nomina.novedad.horas_recargo_nocturno or 0
        hr_nocturno_df = nomina.novedad.horas_recargo_nocturno_dom_fes or 0

        totalh = he_diurnas + he_diurnas_df + he_nocturnas + he_nocturnas_df + hr_diurno_df + hr_nocturo + hr_nocturno_df

        logo_path = "/home/senapress/nomina-sena-django/sena/static/nomina/img/logo/conferencia_motos.png"
        logo_data = None

        # Intentar abrir y codificar la imagen en base64
        try:
            with open(logo_path, "rb") as image_file:
                logo_data = base64.b64encode(image_file.read()).decode("utf-8")
        except IOError:
            print("No se pudo abrir el archivo de imagen.")

        contexto = {
            "nomina": nomina,
            "totalh": totalh,
            "logo_data": logo_data,
        }

        if usuario.correo == 'convergenciamotosgth@gmail.com':
            html_string = render_to_string('nomina/recibo-motos.html', contexto)
        else:
            html_string = render_to_string('nomina/recibo.html', contexto)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Recibo_{nomina.novedad.usuario}_{nomina.fecha_nomina}.pdf"'

        pisa_status = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), response)
        if pisa_status.err:
            return HttpResponse(f'Error al generar el PDF: {pisa_status.err}')
        return response



def descargar_recibo(request, nomina_id):
    nomina = get_object_or_404(Nomina, id=nomina_id)
    nombre_colaborador = nomina.novedad.usuario.nombre
    apellido_colaborador = nomina.novedad.usuario.apellido
    fecha_inicio = nomina.fecha_inicio

    contexto = {
        "nomina": nomina,
    }
    html_string = render_to_string('nomina/recibo.html', contexto)
    response = HttpResponse(content_type='application/pdf')
    response[
        'Content-Disposition'] = f'inline; filename="recibo_{slugify(nombre_colaborador)}{slugify(apellido_colaborador)}_{fecha_inicio}.pdf"'

    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Error al generar el PDF: {pisa_status.err}')
    return response


def parafiscales(request, id):
    if request.session.get("logueo", False):
        nomina = Nomina.objects.get(pk=id)
        print(nomina)
        contexto = {'data': nomina}
        return render(request, 'nomina/parafiscales.html', contexto)


def provisiones(request, id):
    if request.session.get("logueo", False):
        nomina = Nomina.objects.get(pk=id)
        contexto = {'data': nomina}
        return render(request, 'nomina/provisiones.html', contexto)


def liquidacion(request, id):
    if request.session.get("logueo", False):
        usuario = Usuario.objects.get(pk=id)
        nomina = Nomina.objects.filter(novedad__usuario=usuario)
        novedad = Novedad.objects.filter(usuario=usuario).exists()

        if not novedad:
            messages.warning(request, 'Este colaborador necesita tener una nómina para poder ser liquidado.')
            return redirect('nomina:colaboradores')

        contexto = {'usuario': usuario, 'nomina': nomina}
        return render(request, 'nomina/liquidacion/liquidacion.html', contexto)


def liquidacion_archivo(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    nomina = Nomina.objects.get(novedad__usuario=usuario)

    novedades = nomina.novedad_set.all()
    dias_trabajados = sum(novedad.dias_trabajados for novedad in novedades)
    cesantias = (nomina.novedad.salario * dias_trabajados) / 360
    intereses_cesantias = (cesantias * 0.12 * dias_trabajados) / 360
    prima_servicio = (nomina.novedad.salario * dias_trabajados) / 360
    vacaciones = (nomina.novedad.salario * dias_trabajados) / 720
    total_liquidacion = cesantias + intereses_cesantias + prima_servicio + vacaciones
    neto_pagado = total_liquidacion + nomina.total_a_pagar

    contexto = {
        'usuario': usuario,
        'nomina': nomina,
        'dias_trabajados': dias_trabajados,
        "cesantias": cesantias,
        "intereses_cesantias": intereses_cesantias,
        "prima_servicio": prima_servicio,
        "vacaciones": vacaciones,
        "total_liquidacion": total_liquidacion,
        "neto_pagado": neto_pagado,
    }

    return render(request, 'nomina/liquidacion/liquidacion-archivo.html', contexto)


def liquidacion_view(request, id):
    usuario = get_object_or_404(Usuario, id=id)

    # Obtener todas las nóminas relacionadas con el usuario, ordenadas por la fecha fin de la novedad
    nominas = Nomina.objects.filter(novedad__usuario=usuario).order_by('-fecha_nomina')

    if nominas.exists():
        nomina = nominas.first()  # Obtener la nómina más reciente
    else:
        return HttpResponse('No se encontraron nóminas para el usuario.', status=404)

    # Obtener todas las novedades relacionadas con el usuario
    novedades = Novedad.objects.filter(usuario=usuario)
    dias_trabajados = sum(novedad.dias_trabajados for novedad in novedades)

    # Calcular las cesantías y otros valores
    salario = nomina.novedad.usuario.salario
    cesantias = (salario * dias_trabajados) / 360
    intereses_cesantias = (cesantias * 0.12 * dias_trabajados) / 360
    prima_servicio = (salario * dias_trabajados) / 360
    vacaciones = (salario * dias_trabajados) / 720
    total_liquidacion = cesantias + intereses_cesantias + prima_servicio + vacaciones
    neto_pagado = total_liquidacion + nomina.total_a_pagar()

    contexto = {
        'usuario': usuario,
        'nomina': nomina,
        'dias_trabajados': dias_trabajados,
        "cesantias": cesantias,
        "intereses_cesantias": intereses_cesantias,
        "prima_servicio": prima_servicio,
        "vacaciones": vacaciones,
        "total_liquidacion": total_liquidacion,
        "neto_pagado": neto_pagado,
    }

    html_string = render_to_string('nomina/liquidacion/liquidacion-archivo.html', contexto)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Liquidacion_{usuario.nombre}_{usuario.apellido}.pdf"'

    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Error al generar el PDF: {pisa_status.err}')
    return response


def liquidar_colaborador(request, id):
    usuario = Usuario.objects.get(pk=id)
    nominas = Nomina.objects.filter(novedad__usuario=usuario).order_by('-fecha_nomina')

    if nominas.exists():
        nomina = nominas.first()  # Obtener la nómina más reciente
    else:
        return HttpResponse('No se encontraron nóminas para el usuario.', status=404)

    salario = nomina.novedad.usuario.salario
    dias_incapacidad = nomina.novedad.dias_incapacidad or 0
    dias_trabajados = nomina.novedad.dias_trabajados
    horas_extras_diurnas = nomina.novedad.horas_extras_diurnas or 0
    horas_extras_diurnas_dom_fes = nomina.novedad.horas_extras_diurnas_dom_fes or 0
    horas_extras_nocturnas = nomina.novedad.horas_extras_nocturnas or 0
    horas_extras_nocturnas_dom_fes = nomina.novedad.horas_extras_nocturnas_dom_fes or 0
    horas_recargo_nocturno = nomina.novedad.horas_recargo_nocturno or 0
    horas_recargo_nocturno_dom_fes = nomina.novedad.horas_recargo_nocturno_dom_fes or 0
    horas_recargo_diurno_dom_fes = nomina.novedad.horas_recargo_diurno_dom_fes or 0
    comisiones = nomina.novedad.comisiones or 0
    comisiones_porcentaje = nomina.novedad.comisiones_porcentaje or None
    bonificaciones = nomina.novedad.bonificaciones or 0
    embargos_judiciales = nomina.novedad.embargos_judiciales or 0
    libranzas = nomina.novedad.libranzas or 0
    cooperativas = nomina.novedad.cooperativas or 0
    otros = nomina.novedad.otros or 0

    try:
        nov = Novedad(
            usuario=usuario,
            salario=salario,
            dias_incapacidad=dias_incapacidad,
            dias_trabajados=dias_trabajados,
            horas_extras_diurnas=horas_extras_diurnas,
            horas_extras_diurnas_dom_fes=horas_extras_diurnas_dom_fes,
            horas_extras_nocturnas=horas_extras_nocturnas,
            horas_extras_nocturnas_dom_fes=horas_extras_nocturnas_dom_fes,
            horas_recargo_nocturno=horas_recargo_nocturno,
            horas_recargo_nocturno_dom_fes=horas_recargo_nocturno_dom_fes,
            horas_recargo_diurno_dom_fes=horas_recargo_diurno_dom_fes,
            comisiones=comisiones,
            comisiones_porcentaje=comisiones_porcentaje,
            bonificaciones=bonificaciones,
            embargos_judiciales=embargos_judiciales,
            libranzas=libranzas,
            cooperativas=cooperativas,
            otros=otros,
        )
        nov.save()
        messages.success(request, "Colaborador liquidado correctamente!!")
    except Exception as e:
        messages.error(request, f"Error. {e}")

    try:
        q = Usuario.objects.get(pk=id)
        q.nombre = usuario.nombre
        q.apellido = usuario.apellido
        q.correo = usuario.correo
        q.contrasena = usuario.contrasena
        q.rol = usuario.rol
        q.foto = usuario.foto
        q.cargo = usuario.cargo
        q.salario = usuario.salario
        q.riesgo = usuario.riesgo
        q.fecha_fin_contrato = timezone.now()
        q.tipo_contrato = usuario.tipo_contrato
        q.activo = False
        q.fecha_retiro = timezone.now()
        q.motivo_retiro = 'Retiro voluntario'
        q.save()
        messages.success(request, "Colaborador liquidado correctamente!!")
    except Exception as e:
        messages.error(request, f"Error. {e}")

    return HttpResponseRedirect(reverse("nomina:colaboradores"))


def descargar_liquidacion(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    nombre_colaborador = usuario.nombre
    apellido_colaborador = usuario.apellido

    contexto = {
        "usuario": usuario,
        # Añadir aquí cualquier otro dato necesario para la liquidación
    }
    html_string = render_to_string('nomina/liquidacion/liquidacion-archivo.html', contexto)
    response = HttpResponse(content_type='application/pdf')
    response[
        'Content-Disposition'] = f'attachment; filename="liquidacion_{slugify(nombre_colaborador)}_{slugify(apellido_colaborador)}.pdf"'

    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Error al generar el PDF: {pisa_status.err}')
    return response


def descargar_excel(request):
    # Ruta del archivo en el servidor
    file_path = os.path.join(settings.BASE_DIR, 'static', 'nomina', 'files', 'formato_colaborador.xlsm')
    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=formato_colaborador.xlsm'
        return response


def export_nominas_to_excel(request, fecha_nomina):
    # Convertir la fecha recibida en un objeto de fecha
    try:
        fecha_nomina = datetime.strptime(fecha_nomina, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Fecha inválida", status=400)

    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Nóminas"

    # Encabezados de columna
    headers = [
        'Cedula', 'Nombre', 'Apellido', 'Correo', 'Rol', 'Cargo', 'Salario', 'Fecha Ingreso', 'Riesgo',
        'Tipo Contrato', 'Fecha Fin Contrato', 'Activo', 'Fecha Retiro', 'Motivo Retiro',
        'Dias Incapacidad', 'Dias Trabajados', 'Horas Extras Diurnas', 'Horas Extras Diurnas Dom Fes',
        'Horas Extras Nocturnas', 'Horas Extras Nocturnas Dom Fes', 'Horas Recargo Nocturno',
        'Horas Recargo Nocturno Dom Fes', 'Horas Recargo Diurno Dom Fes', 'Comisiones', 'Comisiones Porcentaje',
        'Bonificaciones', 'Embargos Judiciales', 'Libranzas', 'Cooperativas', 'Otros',
        'Salario', 'Comisiones', 'Horas Extras Diurnas ',
        'Horas Extras Diurnas Dom Fes', 'Horas Extras Nocturnas',
        'Horas Extras Nocturnas Dom Fes', 'Horas Recargo Nocturno',
        'Horas Recargo Nocturnos Dom Fes', 'Horas Recargo Diurno Dom Fes',
        'Auxilio Transporte', 'Valor Incapacidad', 'Total Devengado',
        'Salud', 'Pensión', 'FSP', 'Total Deducción',
        'Total a Pagar', 'IBC', 'Riesgo', 'Salud', 'Pensión', 'ARL', 'SENA', 'ICBF', 'Caja Compensación',
        'Cesantías', 'Intereses Cesantías', 'Primas de Servicio', 'Vacaciones', 'Fecha Nómina'
    ]
    ws.append(headers)

    # Obtener las nóminas con la fecha especificada
    nominas = Nomina.objects.select_related('novedad__usuario').prefetch_related('devengado', 'deduccion').filter(
        fecha_nomina=fecha_nomina)

    for nomina in nominas:
        usuario = nomina.novedad.usuario
        novedad = nomina.novedad
        devengado = nomina.devengado
        deduccion = nomina.deduccion

        row = [
            usuario.cedula,
            usuario.nombre,
            usuario.apellido,
            usuario.correo,
            usuario.get_rol_display(),  # Para obtener el texto del rol
            usuario.cargo,
            usuario.salario,
            usuario.fecha_ingreso.strftime('%Y-%m-%d'),
            usuario.riesgo,
            usuario.tipo_contrato,
            usuario.fecha_fin_contrato.strftime('%Y-%m-%d') if usuario.fecha_fin_contrato else '',
            usuario.activo,
            usuario.fecha_retiro.strftime('%Y-%m-%d') if usuario.fecha_retiro else '',
            usuario.motivo_retiro,
            novedad.dias_incapacidad,
            novedad.dias_trabajados,
            novedad.horas_extras_diurnas,
            novedad.horas_extras_diurnas_dom_fes,
            novedad.horas_extras_nocturnas,
            novedad.horas_extras_nocturnas_dom_fes,
            novedad.horas_recargo_nocturno,
            novedad.horas_recargo_nocturno_dom_fes,
            novedad.horas_recargo_diurno_dom_fes,
            novedad.comisiones,
            novedad.comisiones_porcentaje,
            novedad.bonificaciones,
            novedad.embargos_judiciales,
            novedad.libranzas,
            novedad.cooperativas,
            novedad.otros,
            devengado.novedad_salario,
            devengado.novedad_comisiones,
            devengado.valor_horas_extras_diurnas(),
            devengado.valor_horas_extras_diurnas_dom_fes(),
            devengado.valor_horas_extras_nocturnas(),
            devengado.valor_horas_extras_nocturnas_dom_fes(),
            devengado.valor_horas_recargo_nocturno(),
            devengado.valor_horas_recargo_nocturno_dom_fes(),
            devengado.valor_horas_recargo_diurno_dom_fes(),
            devengado.auxilio_transporte(),
            devengado.valor_incapacidad(),
            devengado.total_devengado(),
            deduccion.salud,
            deduccion.pension,
            deduccion.fsp,
            deduccion.total_deduccion(),
            nomina.total_a_pagar(),
            nomina.ibc,
            nomina.riesgo,
            nomina.salud,
            nomina.pension,
            nomina.arl,
            nomina.sena,
            nomina.icbf,
            nomina.caja_compensacion,
            nomina.cesantias,
            nomina.intereses_cesantias,
            nomina.primas_servicio,
            nomina.vacaciones,
            nomina.fecha_nomina.strftime('%Y-%m-%d')
        ]
        ws.append(row)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"nominas_{fecha_nomina}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

def recuperar_contrasena(request):
    if request.method == "POST":
        correo = request.POST.get("correo_recuperar")
        try:
            # Buscar al colaborador por correo
            colaborador = Usuario.objects.get(correo=correo)

            # Preparar el mensaje con la contraseña del usuario
            destinatario = correo
            mensaje = f"""
                <h1 style='color:blue;'>¡Hola, <strong>{colaborador.nombre}</strong>!</h1>
                <p>Tu contraseña fue recuperada exitosamente, esta es tu contraseña: {colaborador.contrasena}</p>
                <a href='https://senapress.pythonanywhere.com'>Ingresa desde este link</a>
            """

            # Intentar enviar el correo
            try:
                msg = EmailMessage("PRESS - Recuperación de contraseña", mensaje, settings.EMAIL_HOST_USER, [destinatario])
                msg.content_subtype = "html"  # Habilitar HTML en el mensaje
                msg.send()
                messages.success(request, "¡El correo con tu contraseña fue enviado exitosamente!")
            except BadHeaderError:
                return HttpResponse("Se encontró un encabezado inválido.")
            except Exception as e:
                return HttpResponse(f"Error al enviar el correo: {e}")

        except Usuario.DoesNotExist:
            # Si el correo no está registrado
            messages.warning(request, "El correo ingresado no coincide con ningún usuario.")
        except Exception as e:
            # En caso de un error inesperado
            messages.error(request, f"Error: {e}")

        return redirect('nomina:login')

    return render(request, 'nomina/login.html')




# -------------------- Modelos --------------------------

def colaboradores(request):
    if request.session.get("logueo", False):
        # Obtén el ID del administrador logueado
        administrador_id = request.session["logueo"]["id"]

        # Filtra los colaboradores creados por el administrador logueado
        q = Usuario.objects.filter(rol=2, activo=True, creado_por_id=administrador_id)
        q2 = Usuario.objects.filter(rol=2, activo=False, creado_por_id=administrador_id)
        admins = Usuario.objects.filter(rol=1, activo=True)
        roles = Usuario.ROLES

        contexto = {
            "data": q,
            'retirados': q2,
            'ROLES': roles,
            "admins": admins
        }

        return render(request, 'nomina/colaborador/colaboradores.html', contexto)
    else:
        messages.warning(request, "Debe iniciar sesión para ver esta página.")
        return HttpResponseRedirect(reverse("nomina:login"))


def prueba(request):
    if request.session.get("logueo", False):
        # Abre el archivo Excel con xlwings
        wb = xw.Book('nomina/excel_files/excel_copia.xlsm')
        sheet_empleados = wb.sheets['EMPLEADOS']
        sheet_datos = wb.sheets['DATOS']

        # Lee los datos de empleados (nombre, cargo) de la hoja EMPLEADOS
        empleados = []
        for row in range(3, 13):  # Filas de 3 a 12
            cedula = sheet_empleados[f'A{row}'].value
            nombre = sheet_empleados[f'B{row}'].value
            cargo = sheet_empleados[f'C{row}'].value

            # Busca el salario correspondiente al cargo en la hoja DATOS
            salarios_rango = sheet_datos.range('B35:C74').value  # Lee todos los cargos y salarios
            salario = 0
            for cargo_datos, salario_datos in salarios_rango:
                if cargo == cargo_datos:
                    salario = salario_datos
                    break

            empleados.append({
                'cedula': cedula,
                'nombre': nombre,
                'cargo': cargo,
                'salario': salario,
            })

        # Asegúrate de actualizar las fórmulas y cálculos
        wb.app.calculate()

        # Cierra el libro sin guardar (opcional si no deseas guardar cambios)
        wb.close()

        contexto = {'empleados': empleados}
        return render(request, 'nomina/prueba.html', contexto)
    else:
        messages.warning(request, "Debe iniciar sesión para ver esta página.")
        return redirect('nomina:login')


def colaborador_guardar(request):
    if request.method == "POST":
        # Obtén los datos del formulario
        nombre = request.POST.get('nombre')
        apellido = request.POST.get("apellido")
        cedula = request.POST.get('cedula')
        correo = request.POST.get('correo')
        contrasena = cedula  # Cambia esto para usar un hash si es necesario
        rol = 2
        foto = request.FILES.get("foto")
        cargo = request.POST.get("cargo")
        salario = request.POST.get("salario")
        riesgo = request.POST.get("riesgo")
        tipo_contrato = request.POST.get("tipo_contrato") or None
        fecha_fin_contrato = request.POST.get("fecha_fin_contrato") or None
        fecha_retiro = request.POST.get("fecha_retiro") or None
        motivo_retiro = request.POST.get("motivo_retiro") or None

        # Asegúrate de obtener el ID del administrador logueado
        administrador_id = request.session.get("logueo", {}).get("id")

        if administrador_id is None:
            messages.error(request, "No se puede determinar quién creó el colaborador.")
            return HttpResponseRedirect(reverse("nomina:colaboradores"))

        try:
            usu = Usuario(
                nombre=nombre,
                apellido=apellido,
                correo=correo,
                cedula=cedula,
                contrasena=contrasena,
                rol=rol,
                foto=foto,
                cargo=cargo,
                salario=salario,
                riesgo=riesgo,
                tipo_contrato=tipo_contrato,
                fecha_fin_contrato=fecha_fin_contrato,
                fecha_retiro=fecha_retiro,
                motivo_retiro=motivo_retiro,
                creado_por_id=administrador_id  # Aquí se guarda el ID del administrador
            )
            usu.save()
            # Resto del código para enviar correo y manejar mensajes

            destinatario = usu.correo

            mensaje = f"""
                <h1 style='color:blue;'>¡Ya eres parte de nuestra familia <strong>PRESS</strong>!</h1>
                <p>Tu registro fue exitoso. Ya puedes acceder a todas nuestras funciones ingresando con tu correo electrónico y esta contraseña: {usu.contrasena}. Puedes cambiarla cuando desees.</p>
                <a href='https://senapress.pythonanywhere.com'>Ingresa desde este link</a>
            """

            try:
                msg = EmailMessage("PRESS", mensaje, settings.EMAIL_HOST_USER, [destinatario])
                msg.content_subtype = "html"  # Habilitar html
                msg.send()
            except BadHeaderError:
                return HttpResponse("Invalid header found.")
            except Exception as e:
                return HttpResponse(f"Error: {e}")
            messages.success(request, "Guardado correctamente!!")
        except Exception as e:
            messages.error(request, f"Error. {e}")
        return HttpResponseRedirect(reverse("nomina:colaboradores", args=()))
    else:
        messages.warning(request, "No se enviarion datos...")
        return HttpResponseRedirect(reverse("nomina:colaboradores", args=()))


def colaborador_guardar_excel(request):
    if request.method == "POST":
        archivo_excel = request.FILES.get('archivo_excel')
        if archivo_excel:
            try:
                wb = openpyxl.load_workbook(archivo_excel)
                hoja = wb['COLABORADORES']
                if hoja:
                    # Obtener el ID del administrador logueado
                    administrador_id = request.session.get("logueo", {}).get("id")

                    if administrador_id is None:
                        messages.error(request, "No se puede determinar quién creó los colaboradores.")
                        return redirect('nomina:colaboradores')

                    for fila in hoja.iter_rows(min_row=4, max_row=hoja.max_row, values_only=True):
                        if len(fila) < 12:
                            messages.warning(request, "Fila de COLABORADORES con número incorrecto de columnas.")
                            continue

                        cedula, nombre, apellido, correo, cargo, salario, fecha_ingreso, riesgo, tipo_contrato, fecha_fin_contrato, fecha_retiro, motivo_retiro = fila[:12]

                        if not cedula or not nombre or not apellido:
                            continue

                        # Crear o actualizar un colaborador
                        col, creado = Usuario.objects.update_or_create(
                            cedula=cedula,
                            defaults={
                                'nombre': nombre,
                                'apellido': apellido,
                                'correo': correo,
                                'contrasena': cedula,  # Considera usar una contraseña encriptada
                                'cargo': cargo,
                                'salario': salario,
                                'fecha_ingreso': fecha_ingreso,
                                'riesgo': riesgo,
                                'tipo_contrato': tipo_contrato,
                                'fecha_fin_contrato': fecha_fin_contrato,
                                'fecha_retiro': fecha_retiro,
                                'motivo_retiro': motivo_retiro,
                                'rol': 2,  # Asignar rol de colaborador
                                'creado_por_id': administrador_id  # Guardar el ID del administrador que creó
                            }
                        )

                        hoja_novedades = wb['NOVEDADES']  # Accede a la hoja de novedades
                        for fila_novedad in hoja_novedades.iter_rows(min_row=4, max_row=hoja_novedades.max_row, values_only=True):
                            # Aseguramos que la fila de novedades tenga al menos 20 columnas
                            if len(fila_novedad) < 20:
                                messages.warning(request, "Fila de NOVEDADES con número incorrecto de columnas.")
                                continue

                            # Solo tomamos las primeras 20 columnas
                            nov_cedula, incapacidad, dias_trabajados, perm_remunerado, perm_no_remunerado, sin_justa_causa, he_diurna, he_diurna_dom_fes, he_nocturna, he_nocturna_dom_fes, rec_nocturno, rec_diurno_dom_fes, rec_nocturno_dom_fes, comision_venta, comision_porcentaje, bonificaciones, embargos, libranzas, cooperativas, otros = fila_novedad[:20]

                            # Verifica si la cédula de la novedad coincide con el colaborador
                            if nov_cedula == cedula:
                                # Comprueba si la novedad ya existe para evitar duplicados
                                novedad_existente = Novedad.objects.filter(usuario=col).exists()

                                if not novedad_existente:
                                    # Crear una nueva novedad para el colaborador si no existe
                                    nov = Novedad(
                                        usuario=col,
                                        dias_incapacidad=incapacidad,
                                        dias_trabajados=dias_trabajados,
                                        perm_remunerado=perm_remunerado,
                                        perm_no_remunerado=perm_no_remunerado,
                                        sin_justa_causa=sin_justa_causa,
                                        horas_extras_diurnas=he_diurna,
                                        horas_extras_diurnas_dom_fes=he_diurna_dom_fes,
                                        horas_extras_nocturnas=he_nocturna,
                                        horas_extras_nocturnas_dom_fes=he_nocturna_dom_fes,
                                        horas_recargo_nocturno=rec_nocturno,
                                        horas_recargo_diurno_dom_fes=rec_diurno_dom_fes,
                                        horas_recargo_nocturno_dom_fes=rec_nocturno_dom_fes,
                                        comisiones=comision_venta,
                                        comisiones_porcentaje=comision_porcentaje,
                                        bonificaciones=bonificaciones,
                                        embargos_judiciales=embargos,
                                        libranzas=libranzas,
                                        cooperativas=cooperativas,
                                        otros=otros,
                                    )
                                    nov.save()

                    messages.success(request, "Colaboradores y novedades importados con éxito")
                    return redirect('nomina:colaboradores')

                else:
                    messages.warning(request, "El archivo de Excel no cumple con el formato requerido.")
            except Exception as e:
                messages.error(request, f"Error al cargar el archivo Excel: {str(e)}")
                return redirect('nomina:colaboradores')

    return render(request, 'nomina/colaborador_form.html')


def colaborador_editar(request, id):
    if request.session.get("logueo", False):
        usuario = Usuario.objects.get(pk=id)
        nombre = request.POST.get('nombre_editar')
        apellido = request.POST.get("apellido_editar")
        cedula = request.POST.get('cedula_editar')
        correo = request.POST.get('correo_editar')
        contrasena = request.POST.get('contrasena_editar')
        confirmar_contrasena = request.POST.get('confirmar_contrasena_editar')
        rol = usuario.rol
        foto = request.FILES.get("foto_editar")
        cargo = request.POST.get("cargo_editar")
        salario = request.POST.get("salario_editar")
        riesgo = request.POST.get("riesgo_editar") or usuario.riesgo
        tipo_contrato = request.POST.get("tipo_contrato_editar") or None
        fecha_fin_contrato = request.POST.get("fecha_fin_contrato_editar") or None
        fecha_retiro = request.POST.get("fecha_retiro_editar") or None
        motivo_retiro = request.POST.get("motivo_retiro_editar") or None

        try:
            q = Usuario.objects.get(pk=id)
            q.nombre = nombre
            q.apellido = apellido
            q.cedula = cedula
            q.correo = correo
            if contrasena == confirmar_contrasena:
                q.contrasena = contrasena
            else:
                messages.error(request, f"Las contraseñas no coinciden...")
            q.rol = rol
            if foto:
                q.foto = foto
            q.cargo = cargo
            q.salario = salario
            q.riesgo = riesgo
            q.tipo_contrato = tipo_contrato
            q.fecha_fin_contrato = fecha_fin_contrato
            q.fecha_retiro = fecha_retiro
            q.motivo_retiro = motivo_retiro
            q.save()
            messages.success(request, "Actualizado correctamente!!")
        except Exception as e:
            messages.error(request, f"Error. {e}")

        return redirect('nomina:colaboradores')


def colaborador_eliminar(request, id):
    try:
        q = Usuario.objects.get(pk=id)
        q.delete()
        messages.success(request, "Eliminado correctamente!!")
    except Exception as e:
        messages.error(request, f"Error. {e}")
    return HttpResponseRedirect(reverse("nomina:colaboradores", args=()))


def novedades_nomina(request):
    if request.session.get("logueo", False):
        administrador_id = request.session["logueo"]["id"]
        novedades = Novedad.objects.filter(usuario__creado_por_id=administrador_id)
        novedades_admin = Novedad.objects.filter(usuario__id=administrador_id)
        usuarios = Usuario.objects.filter(rol=2, creado_por_id=administrador_id)
        contexto = {
            "novedades": novedades,
            "novedades_admin": novedades_admin,
            "usuarios": usuarios
        }
        return render(request, 'nomina/novedad/novedades_nomina.html', contexto)
    else:
        messages.warning(request, "Debe iniciar sesión para ver esta página.")
        return HttpResponseRedirect(reverse("nomina:login"))


def novedad_guardar(request):
    if request.method == "POST":
        usuario_id = request.POST.get("usuario")
        usuario = Usuario.objects.get(id=usuario_id)
        dias_incapacidad = request.POST.get("dias_incapacidad") or 0
        dias_trabajados = request.POST.get("dias_trabajados") or 15
        perm_remunerado = request.POST.get("perm_remunerado") or None
        perm_no_remunerado = request.POST.get("perm_no_remunerado") or None
        sin_justa_causa = request.POST.get("sin_justa_causa") or None
        horas_extras_diurnas = request.POST.get("horas_extras_diurnas") or 0
        horas_extras_diurnas_dom_fes = request.POST.get("horas_extras_diurnas_dom_fes") or 0
        horas_extras_nocturnas = request.POST.get("horas_extras_nocturnas") or 0
        horas_extras_nocturnas_dom_fes = request.POST.get("horas_extras_nocturnas_dom_fes") or 0
        horas_recargo_nocturno = request.POST.get("horas_recargo_nocturno") or 0
        horas_recargo_nocturno_dom_fes = request.POST.get("horas_recargo_nocturno_dom_fes") or 0
        horas_recargo_diurno_dom_fes = request.POST.get("horas_recargo_diurno_dom_fes") or 0
        comisiones = request.POST.get("comisiones") or 0
        comisiones_porcentaje = request.POST.get("comisiones_porcentaje") or None
        bonificaciones = request.POST.get("bonificaciones") or 0
        embargos_judiciales = request.POST.get("embargos_judiciales") or 0
        libranzas = request.POST.get("libranzas") or 0
        cooperativas = request.POST.get("libranzas") or 0
        otros = request.POST.get("otros") or 0

        try:
            nov = Novedad(
                usuario=usuario,
                dias_incapacidad=dias_incapacidad,
                dias_trabajados=dias_trabajados,
                perm_remunerado=perm_remunerado,
                perm_no_remunerado=perm_no_remunerado,
                sin_justa_causa=sin_justa_causa,
                horas_extras_diurnas=horas_extras_diurnas,
                horas_extras_diurnas_dom_fes=horas_extras_diurnas_dom_fes,
                horas_extras_nocturnas=horas_extras_nocturnas,
                horas_extras_nocturnas_dom_fes=horas_extras_nocturnas_dom_fes,
                horas_recargo_nocturno=horas_recargo_nocturno,
                horas_recargo_nocturno_dom_fes=horas_recargo_nocturno_dom_fes,
                horas_recargo_diurno_dom_fes=horas_recargo_diurno_dom_fes,
                comisiones=comisiones,
                comisiones_porcentaje=comisiones_porcentaje,
                bonificaciones=bonificaciones,
                embargos_judiciales=embargos_judiciales,
                libranzas=libranzas,
                cooperativas=cooperativas,
                otros=otros,
            )
            nov.save()
            messages.success(request, "Guardado correctamente!!")
        except Exception as e:
            messages.error(request, f"Error. {e}")

        return HttpResponseRedirect(reverse("nomina:novedades_nomina"))
    else:
        messages.warning(request, "No se enviaron datos...")
        return HttpResponseRedirect(reverse("nomina:novedades_nomina"))


def novedad_admin_guardar(request, id):
    try:
        # Verifica si el usuario ya tiene una novedad
        novedad_existente = Novedad.objects.filter(usuario__id=id).exists()

        if novedad_existente:
            # Si ya existe una novedad, muestra un mensaje de error
            messages.error(request, "Este usuario ya tiene una novedad registrada.")
            return HttpResponseRedirect(reverse("nomina:colaboradores"))
        else:
            # Si no existe, crea una nueva novedad
            nov = Novedad(
                usuario=Usuario.objects.get(id=id),
                dias_incapacidad=0,
                dias_trabajados=15,
                perm_remunerado=None,
                perm_no_remunerado=None,
                sin_justa_causa=None,
                horas_extras_diurnas=0,
                horas_extras_diurnas_dom_fes=0,
                horas_extras_nocturnas=0,
                horas_extras_nocturnas_dom_fes=0,
                horas_recargo_nocturno=0,
                horas_recargo_nocturno_dom_fes=0,
                horas_recargo_diurno_dom_fes=0,
                comisiones=0,
                comisiones_porcentaje=None,
                bonificaciones=0,
                embargos_judiciales=0,
                libranzas=0,
                cooperativas=0,
                otros=0
            )
            nov.save()
            # Muestra un mensaje de éxito
            messages.success(request, "Novedad guardada correctamente!!")
            return HttpResponseRedirect(reverse("nomina:novedades_nomina"))
    except Exception as e:
        # Si ocurre algún error, muestra un mensaje de error
        messages.error(request, f"Error al guardar la novedad: {e}")
        return HttpResponseRedirect(reverse("nomina:colaboradores"))
    # Redirige de nuevo a la página de novedades



def novedad_editar(request, id):
    if request.session.get("logueo", False):
        novedad = get_object_or_404(Novedad, id=id)
        usuario_id = request.POST.get("usuario_editar")
        usuario = Usuario.objects.get(id=usuario_id)
        dias_incapacidad = request.POST.get("dias_incapacidad_editar") or novedad.dias_incapacidad
        dias_trabajados = request.POST.get("dias_trabajados_editar") or novedad.dias_trabajados
        perm_remunerado = request.POST.get("perm_remunerado_editar") or novedad.perm_remunerado
        perm_no_remunerado = request.POST.get("perm_no_remunerado_editar") or novedad.perm_no_remunerado
        sin_justa_causa = request.POST.get("sin_justa_causa_editar") or novedad.sin_justa_causa
        horas_extras_diurnas = request.POST.get("horas_extras_diurnas_editar") or novedad.horas_extras_diurnas
        horas_extras_diurnas_dom_fes = request.POST.get(
            "horas_extras_diurnas_dom_fes_editar") or novedad.horas_extras_diurnas_dom_fes
        horas_extras_nocturnas = request.POST.get("horas_extras_nocturnas_editar") or novedad.horas_extras_nocturnas
        horas_extras_nocturnas_dom_fes = request.POST.get(
            "horas_extras_nocturnas_dom_fes_editar") or novedad.horas_extras_nocturnas_dom_fes
        horas_recargo_nocturno = request.POST.get("horas_recargo_nocturno_editar") or novedad.horas_recargo_nocturno
        horas_recargo_nocturno_dom_fes = request.POST.get(
            "horas_recargo_nocturno_dom_fes_editar") or novedad.horas_recargo_nocturno_dom_fes
        horas_recargo_diurno_dom_fes = request.POST.get(
            "horas_recargo_diurno_dom_fes_editar") or novedad.horas_recargo_diurno_dom_fes
        comisiones = request.POST.get("comisiones_editar") or novedad.comisiones
        comisiones_porcentaje = request.POST.get("comisiones_porcentaje_editar") or novedad.comisiones_porcentaje
        bonificaciones = request.POST.get("bonificaciones_editar") or novedad.bonificaciones
        embargos_judiciales = request.POST.get("embargos_judiciales_editar") or novedad.embargos_judiciales
        libranzas = request.POST.get("libranzas_editar") or novedad.libranzas
        cooperativas = request.POST.get("cooperativas_editar") or novedad.cooperativas
        otros = request.POST.get("otros_editar") or novedad.otros

        try:
            # Actualizar la novedad
            q = Novedad.objects.get(pk=id)
            q.usuario = usuario
            q.dias_incapacidad = dias_incapacidad
            q.dias_trabajados = dias_trabajados
            q.perm_remunerado = perm_remunerado
            q.perm_no_remunerado = perm_no_remunerado
            q.sin_justa_causa = sin_justa_causa
            q.horas_extras_diurnas = horas_extras_diurnas
            q.horas_extras_diurnas_dom_fes = horas_extras_diurnas_dom_fes
            q.horas_extras_nocturnas = horas_extras_nocturnas
            q.horas_extras_nocturnas_dom_fes = horas_extras_nocturnas_dom_fes
            q.horas_recargo_nocturno = horas_recargo_nocturno
            q.horas_recargo_nocturno_dom_fes = horas_recargo_nocturno_dom_fes
            q.horas_recargo_diurno_dom_fes = horas_recargo_diurno_dom_fes
            q.comisiones = comisiones
            q.comisiones_porcentaje = comisiones_porcentaje
            q.bonificaciones = bonificaciones
            q.embargos_judiciales = embargos_judiciales
            q.libranzas = libranzas
            q.cooperativas = cooperativas
            q.otros = otros
            q.fecha_ultima_actualizacion = timezone.now()  # Actualizamos la fecha de última actualización
            q.save()

            messages.success(request, f"La novedad de {usuario} fue actualizada!")
        except Exception as e:
            messages.error(request, f"Error. {e}")

        return redirect('nomina:reportar_novedad')

def editar_porcentaje_incapacidad(request):
    if request.session.get("logueo", False):
        novedades = Novedad.objects.all()
        incapacidad_porcentaje = request.POST.get("incapacidad_porcentaje_editar")

        if incapacidad_porcentaje:
            try:
                porcentaje_float = float(incapacidad_porcentaje) / 100
                # Recorre todas las novedades y actualiza el campo
                for novedad in novedades:
                    novedad.incapacidad_porcentaje = porcentaje_float
                    novedad.save()
                messages.success(request, "¡Las novedades fueron actualizadas!")
            except Exception as e:
                messages.error(request, f"Error al actualizar. {e}")

        return HttpResponseRedirect(reverse("nomina:index"))




def novedad_eliminar(request, id):
    try:
        q = Novedad.objects.get(pk=id)
        q.delete()
        messages.success(request, "Eliminado correctamente!!")
    except Exception as e:
        messages.error(request, f"Error. {e}")
    return HttpResponseRedirect(reverse("nomina:novedades_nomina", args=()))


def reportar_novedad(request):
    if request.session.get("logueo", False):
        administrador_id = request.session["logueo"]["id"]
        novedades = Novedad.objects.filter(usuario__creado_por_id=administrador_id)
        novedades_admin = Novedad.objects.filter(usuario__id=administrador_id)
        usuarios = Usuario.objects.filter(rol=2, creado_por_id=administrador_id)
        contexto = {
            "novedades": novedades,
            "novedades_admin": novedades_admin,
            "usuarios": usuarios
        }
        return render(request, 'nomina/novedad/reportar_novedad.html', contexto)
    else:
        messages.warning(request, "Debe iniciar sesión para ver esta página.")
        return HttpResponseRedirect(reverse("nomina:login"))


def actualizar_novedades(request):
    if request.method == 'POST':
        # Obtener la fecha actual
        fecha_actual = timezone.now()

        # Actualizar la fecha_ultima_actualizacion de todas las novedades
        novedades = Novedad.objects.all()
        novedades.update(fecha_ultima_actualizacion=fecha_actual)

        # Crear devengados, deducciones y nóminas para cada novedad
        for novedad in novedades:
            # Crear el devengado
            devengado = Devengado.objects.create(novedad=novedad)

            # Crear la deducción
            deduccion = Deduccion.objects.create(novedad=novedad, devengado=devengado, retefuente='Valor ejemplo')

            # Crear la nómina
            Nomina.objects.create(
                novedad=novedad,
                devengado=devengado,
                deduccion=deduccion,
                fecha_nomina=novedad.fecha_ultima_actualizacion
            )

        messages.success(request, f"Se reportaron las novedades y nóminas de la fecha: {fecha_actual}")
    return redirect('nomina:novedades_nomina')



def nomina(request):
    if request.session.get("logueo", False):
        usuario_id = request.session["logueo"]["id"]
        try:
            usuario = Usuario.objects.get(id=usuario_id)

            if usuario.rol == 2:
                nominas = Nomina.objects.filter(novedad__usuario=usuario)
                total_a_pagar_usuario = sum(n.total_a_pagar() for n in nominas)
            else:
                # Obtener las nóminas de los colaboradores
                nomina1 = Nomina.objects.filter(novedad__usuario__creado_por_id=usuario_id)

                # Obtener la nómina del propio administrador
                nomina2 = Nomina.objects.filter(novedad__usuario__id=usuario_id)

                # Eliminar duplicados, si la nómina del administrador ya está en nomina1
                if nomina2:
                    nomina1 = nomina1.exclude(id=nomina2[0].id)

                # Combinar las nóminas de los colaboradores con la del administrador
                nominas = list(chain(nomina1, nomina2))
                total_a_pagar_usuario = None

            # Agrupar nóminas por fecha de nómina
            nominas_por_fecha_nomina = {}
            for nomina in nominas:
                fecha_nomina = nomina.fecha_nomina
                if fecha_nomina not in nominas_por_fecha_nomina:
                    nominas_por_fecha_nomina[fecha_nomina] = []
                nominas_por_fecha_nomina[fecha_nomina].append(nomina)

            # Ordenar las fechas (por defecto las fechas se ordenarán en orden ascendente)
            nominas_por_fecha_nomina = dict(sorted(nominas_por_fecha_nomina.items()))

            # Verificar si hay nóminas
            if not nominas_por_fecha_nomina:
                mensaje = "No hay nóminas disponibles."
            else:
                mensaje = None

            contexto = {
                "nominas_por_fecha_nomina": nominas_por_fecha_nomina,
                "total_a_pagar_usuario": total_a_pagar_usuario,
                "mensaje": mensaje
            }
            return render(request, 'nomina/nomina/nomina.html', contexto)
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado")
            return HttpResponseRedirect(reverse("nomina:login"))
    else:
        messages.warning(request, "Debe iniciar sesión para ver esta página.")
        return HttpResponseRedirect(reverse("nomina:login"))


def nomina_listar(request, fecha_nomina):
    if request.session.get("logueo", False):
        usuario_id = request.session["logueo"]["id"]
        usuario = Usuario.objects.get(id=usuario_id)

        if usuario.rol == 1:
            try:
                nomina1 = Nomina.objects.filter(fecha_nomina=fecha_nomina, novedad__usuario__creado_por_id=usuario_id)

                # Obtener la nómina del propio administrador
                nomina2 = Nomina.objects.filter(fecha_nomina=fecha_nomina, novedad__usuario__id=usuario_id)

                # Eliminar duplicados, si la nómina del administrador ya está en nomina1
                if nomina2:
                    nomina1 = nomina1.exclude(id=nomina2[0].id)

                # Combinar las nóminas de los colaboradores con la del administrador
                nominas = list(chain(nomina1, nomina2))

                print(nominas[0].fecha_nomina)

                contexto = {
                    "nominas": nominas,
                    "fecha_ultima_actualizacion": fecha_nomina
                }

                if not nominas:
                    contexto['mensaje'] = "No hay nóminas disponibles para esta fecha."

                return render(request, 'nomina/nomina/nomina-listar.html', contexto)

            except Usuario.DoesNotExist:
                messages.error(request, "Usuario no encontrado")
                return HttpResponseRedirect(reverse("nomina:nomina_listar"))

        if usuario.rol == 2:
            try:
                nominas = Nomina.objects.filter(novedad__usuario=usuario)

                contexto = {
                    "nominas": nominas
                }

                if not nominas:
                    contexto['mensaje'] = "No hay nóminas disponibles para este usuario."

                return render(request, 'nomina/nomina/nomina-listar.html', contexto)
            except Usuario.DoesNotExist:
                messages.error(request, "Usuario no encontrado")
                return HttpResponseRedirect(reverse("nomina:nomina_listar"))
    else:
        messages.warning(request, "Debe iniciar sesión para ver esta página.")
        return HttpResponseRedirect(reverse("nomina:login"))



def usuario_listar(request):
    return render(request, 'nomina/usuario/usuario-listar.html')


def usuario_formulario(request):
    return render(request, 'nomina/usuario/usuario-formulario.html')


def usuario_guardar_imagen(f, nuevo_nombre):
    with open(f"uploads/fotos/{nuevo_nombre}", "wb+") as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def usuario_guardar(request):
    if request.method == "POST":
        id = request.POST.get("id")
        cedula = request.POST.get("cedula")
        nombre = request.POST.get('nombre')
        apellido = request.POST.get("apellido")
        correo = request.POST.get('correo')
        contrasena = request.POST.get('contrasena')
        rol = 1
        foto = request.FILES.get("foto")
        cargo = request.POST.get("cargo") or None
        riesgo = request.POST.get("riesgo") or 0
        fecha_fin_contrato = request.POST.get("fecha_fin_contrato") or None
        tipo_contrato = request.POST.get("tipo_contrato") or None
        fecha_retiro = request.POST.get("fecha_retiro") or None
        motivo_retiro = request.POST.get("motivo_retiro") or None

        if cargo is not None:
            try:
                cargo = int(cargo)  # Convertir el valor de cargo a entero
            except ValueError:
                cargo = None

        if id == "":
            # crear
            try:
                usu = Usuario(
                    cedula=cedula,
                    nombre=nombre,
                    apellido=apellido,
                    correo=correo,
                    contrasena=contrasena,
                    rol=rol,
                    foto=foto,
                    cargo=cargo,
                    riesgo=riesgo,
                    fecha_fin_contrato=fecha_fin_contrato,
                    tipo_contrato=tipo_contrato,
                    fecha_retiro=fecha_retiro,
                    motivo_retiro=motivo_retiro,
                )
                usu.save()
                messages.success(request, "Guardado correctamente!!")
                try:
                    destinatario = usu.correo

                    mensaje = f"""
                        <h1 style='color:blue;'>¡Ya eres parte de nuestra familia <strong>PRESS</strong>!</h1>
                        <p>Tu registro fue exitoso. Ya puedes acceder a todas nuestras funciones ingresando con tu correo electrónico y esta contraseña: {usu.contrasena}. Puedes cambiarla cuando desees.</p>
                        <a href='https://senapress.pythonanywhere.com'>Ingresa desde este link</a>
                    """

                    try:
                        msg = EmailMessage("PRESS", mensaje, settings.EMAIL_HOST_USER, [destinatario])
                        msg.content_subtype = "html"  # Habilitar html
                        msg.send()
                    except BadHeaderError:
                        return HttpResponse("Invalid header found.")
                    except Exception as e:
                        return HttpResponse(f"Error: {e}")

                except Exception as e:
                    print(f"{e}")

            except Exception as e:
                messages.error(request, f"Error. {e}")
        else:
            # actualizar
            try:
                q = Usuario.objects.get(pk=id)
                q.cedula = cedula
                q.nombre = nombre
                q.apellido = apellido
                q.correo = correo
                q.contrasena = contrasena
                q.rol = rol
                q.foto = foto
                q.cargo = cargo
                q.riesgo = riesgo
                q.fecha_fin_contrato = fecha_fin_contrato
                q.tipo_contrato = tipo_contrato
                q.fecha_retiro = fecha_retiro
                q.motivo_retiro = motivo_retiro
                q.save()
                messages.success(request, "Actualizado correctamente!!")
            except Exception as e:
                messages.error(request, f"Error. {e}")

        return HttpResponseRedirect(reverse("nomina:index", args=()))

    else:
        messages.warning(request, "No se enviarion datos...")
        return HttpResponseRedirect(reverse("nomina:usuario_formulario", args=()))
